import streamlit as st
import pandas as pd
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import io

# إعدادات واجهة الموقع
st.set_page_config(page_title="معالج حالات A79 - يوسف", layout="wide")
st.markdown("<h1 style='text-align: center; color: #2F5597;'>📊 منصة معالجة حالات الغش A79</h1>", unsafe_allow_case_尊尊_html=True)
st.write("---")

# تنسيقات الإكسيل
months_fr = {1:"Janvier", 2:"Février", 3:"Mars", 4:"Avril", 5:"Mai", 6:"Juin", 
             7:"Juillet", 8:"Août", 9:"Septembre", 10:"Octobre", 11:"Novembre", 12:"Décembre"}
header_fill = PatternFill(start_color="2F5597", fill_type="solid")
white_font = Font(color="FFFFFF", bold=True)
std_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

# دالة المعالجة الأساسية
def process_files(ventes_files, ca_files):
    all_results = []
    
    # تحويل الملفات المرفوعة إلى قاموس ليسهل الوصول إليها
    ventes_dict = {f.name: f for f in ventes_files}
    ca_dict = {f.name: f for f in ca_files}
    
    for energy in ['BT', 'BP']:
        v_file = next((f for name, f in ventes_dict.items() if energy in name.upper() and 'VENTES' in name.upper()), None)
        c_file = next((f for name, f in ca_dict.items() if energy in name.upper() and 'CA' in name.upper()), None)
        
        if v_file and c_file:
            df_v = pd.read_excel(v_file, skiprows=5)
            df_ca = pd.read_excel(c_file, skiprows=5)
            df_v.columns = [str(c).strip() for c in df_v.columns]
            df_ca.columns = [str(c).strip() for c in df_ca.columns]
            
            col_ev = next((c for c in df_v.columns if 'V\u00c8NEMENT' in c.upper() or 'EVENEMENT' in c.upper()), 'Numéro évènement')
            col_ft = 'Numéro Facture'
            col_ttc = next((c for c in df_ca.columns if 'TTC' in c.upper()), None)
            
            # تنظيف رقم الفاتورة كما طلبت (حذف .0)
            for df in [df_v, df_ca]:
                df[col_ev] = df[col_ev].astype(str).str.strip()
                df[col_ft] = df[col_ft].apply(lambda x: str(x).split('.')[0].strip() if pd.notnull(x) else "")
            
            df_v_a79 = df_v[df_v[col_ev].str.contains("A79", na=False)].copy()
            df_m = pd.merge(df_v_a79, df_ca[[col_ev, col_ft, col_ttc]], on=[col_ev, col_ft], how='inner')
            
            if not df_m.empty:
                df_m['Energy_Type'] = 'ÉLECTRICITÉ' if energy == 'BT' else 'GAZ'
                df_m['Mois_Num'] = pd.to_datetime(df_m['Date'], errors='coerce').dt.month
                df_m['Mtt_TTC'] = pd.to_numeric(df_m[col_ttc], errors='coerce').fillna(0)
                all_results.append(df_m)
                
    if not all_results: return None
    return pd.concat(all_results, ignore_index=True)

# واجهة المستخدم
st.subheader("1️⃣ ارفع الملفات المطلوبة")
col1, col2 = st.columns(2)
with col1:
    v_uploads = st.file_uploader("ملفات Detail_Ventes", accept_multiple_files=True)
with col2:
    c_uploads = st.file_uploader("ملفات Detail_CA_Energie", accept_multiple_files=True)

if st.button("🚀 بدء المعالجة وإنتاج التقرير"):
    if v_uploads and c_uploads:
        final_df = process_files(v_uploads, c_uploads)
        if final_df is not None:
            st.success("✅ تم دمج البيانات بنجاح ومطابقة الفواتير!")
            
            # تحويل البيانات إلى إكسيل في الذاكرة
            output = io.BytesIO()
            with pd.ExcelWriter(output, engine='openpyxl') as writer:
                final_df['Mois_Nom'] = final_df['Mois_Num'].map(months_fr)
                cols = ['Energy_Type', 'Mois_Nom', 'Agence', 'Numéro évènement', 'Numéro Facture', 'Code client', 'Total énergie (Kwh)', 'Mtt_TTC']
                final_df[cols].to_excel(writer, sheet_name="Détails_Fraude", index=False)
                
                # تنسيق بسيط للجمالية
                ws = writer.sheets["Détails_Fraude"]
                for i in range(1, len(cols) + 1):
                    ws.cell(1, i).fill = header_fill
                    ws.cell(1, i).font = white_font
                    ws.column_dimensions[get_column_letter(i)].width = 20
            
            st.download_button(
                label="📥 تحميل التقرير النهائي (Excel)",
                data=output.getvalue(),
                file_name="Rapport_Fraude_Youssef.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            )
        else:
            st.warning("⚠️ لم يتم العثور على حالات مطابقة بين الملفات.")
    else:
        st.error("❌ يرجى رفع الملفات أولاً.")
