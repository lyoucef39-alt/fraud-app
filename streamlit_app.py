import pandas as pd
import glob
import os
import re
import warnings
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

warnings.filterwarnings("ignore", category=UserWarning, module='openpyxl')

# ==========================================
# 1. إعدادات وتنسيقات مشتركة
# ==========================================
fmt_accounting = '_-* # ##0.00_-;-* # ##0.00_-;_-* "-"??_-;_-@_-'
months_fr = ["Janvier", "Février", "Mars", "Avril", "Mai", "Juin", "Juillet", "Août", "Septembre", "Octobre", "Novembre", "Décembre"]
header_fill = PatternFill(start_color="2F5597", fill_type="solid")
mid_fill = PatternFill(start_color="4472C4", fill_type="solid")
total_fill = PatternFill(start_color="D9E1F2", fill_type="solid")
white_bold = Font(color="FFFFFF", bold=True)
std_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

def extract_agency_code(x):
    s = str(x).strip()
    if re.match(r'^\d{3}', s) and "Date" not in s and "génération" not in s:
        return s[:3]
    return None

# ==========================================
# 2. وظائف معالجة البيانات
# ==========================================
def get_data_v1(energy_type):
    pattern = os.path.join('data', f"*Detail_Ventes*_{energy_type}_*.xlsx")
    files = glob.glob(pattern)
    if not files: return pd.DataFrame()
    df = pd.read_excel(files[0], skiprows=5)
    df.columns = [str(c).strip() for c in df.columns]
    val_col = next(c for c in df.columns if any(k in c.lower() for k in ['énergie', 'consom', 'total']))
    ag_col = next(c for c in df.columns if 'AGENCE' in c.upper())
    if energy_type in ['BT', 'BP']:
        df['Mois'] = pd.to_datetime(df['Date'], errors='coerce').dt.month
    else:
        f_col = next(c for c in df.columns if 'FACTURE' in c.upper())
        start = 4 if energy_type == 'HTA' else 5
        df['Mois'] = df[f_col].apply(lambda x: int(str(x)[start:start+2]) if len(str(x)) > start+1 else None)
    df['Val'] = pd.to_numeric(df[val_col], errors='coerce').fillna(0)
    df['Ag_Code'] = df[ag_col].apply(extract_agency_code)
    df['Ag_Full'] = df[ag_col]
    df['Nat'] = df['Nature'].apply(lambda n: 'CYCL' if 'CYCL' in str(n).upper() else ('HC' if 'HC' in str(n).upper() else ('ANNUL' if 'ANNUL' in str(n).upper() else 'OTHER')))
    return df[df['Ag_Code'].notna()]

def get_stats_data_v3(energy_type):
    pattern = os.path.join('data', f"*Detail_Ventes*_{energy_type}_*.xlsx")
    files = glob.glob(pattern)
    if not files: return pd.DataFrame()
    df = pd.read_excel(files[0], skiprows=5)
    df.columns = [str(c).strip() for c in df.columns]
    val_col = next(c for c in df.columns if any(k in c.lower() for k in ['énergie', 'consom', 'total']))
    type_col = next(c for c in df.columns if 'TYPE' in c.upper() or 'CLIENT' in c.upper())
    nat_col = next(c for c in df.columns if 'NATURE' in c.upper())
    df = df[df[nat_col].astype(str).str.contains('EMS CYCL', na=False, case=False)]
    df['Date_Conv'] = pd.to_datetime(df['Date'], errors='coerce')
    df['Mois'] = df['Date_Conv'].dt.month
    df['Annee'] = df['Date_Conv'].dt.year.astype(str).replace('nan', '-')
    df['Cat'] = df[type_col].apply(lambda x: 'AO' if 'AO' in str(x).upper() else ('FSM' if 'FSM' in str(x).upper() else 'OTHER'))
    df['Is_Nulle'] = pd.to_numeric(df[val_col], errors='coerce').fillna(0) == 0
    return df

# ==========================================
# 3. التشغيل الرئيسي
# ==========================================
def run_mega_process():
    print("🚀 جاري معالجة البيانات وتجهيز التقارير النهائية...")
    
    energies = ['BT', 'HTA', 'BP', 'MP']
    data_v1 = {e: get_data_v1(e) for e in energies}
    all_codes = sorted(set().union(*[df['Ag_Code'].unique() for df in data_v1.values() if not df.empty]))
    code_to_name = {}
    for e in energies:
        if not data_v1[e].empty:
            code_to_name.update(data_v1[e].drop_duplicates('Ag_Code').set_index('Ag_Code')['Ag_Full'].to_dict())

    def get_stats_dash(energy, is_bt_bp=True):
        df = data_v1[energy]
        if df.empty: return (pd.Series(0, index=range(1,13)), pd.Series(0, index=range(1,13))) if is_bt_bp else pd.Series(0, index=range(1,13))
        df['Val_C'] = df.apply(lambda r: -r['Val'] if r['Nat'] == 'ANNUL' else r['Val'], axis=1)
        if is_bt_bp:
            type_col = next((c for c in df.columns if 'TYPE' in c.upper() or 'CLIENT' in c.upper()), None)
            df['Cat'] = df[type_col].apply(lambda x: 'AO' if 'AO' in str(x).upper() else ('FSM' if 'FSM' in str(x).upper() else 'OTHER'))
            summary = df.groupby(['Mois', 'Cat'], observed=False)['Val_C'].sum().unstack(fill_value=0).reindex(range(1, 13), fill_value=0)
            return summary.get('AO', pd.Series(0, index=range(1,13))), summary.get('FSM', pd.Series(0, index=range(1,13)))
        return df.groupby('Mois')['Val_C'].sum().reindex(range(1, 13), fill_value=0)

    bt_ao, bt_fsm = get_stats_dash('BT'); hta_v = get_stats_dash('HTA', False)
    bp_ao, bp_fsm = get_stats_dash('BP'); mp_v = get_stats_dash('MP', False)

    df_dash = pd.DataFrame(index=months_fr)
    df_dash['Elec_AO'], df_dash['Elec_FSM'] = bt_ao.values, bt_fsm.values
    df_dash['BT (Total)'] = df_dash['Elec_AO'] + df_dash['Elec_FSM']
    df_dash['MT (HTA)'] = hta_v.values
    df_dash['Total Elec'] = df_dash['BT (Total)'] + df_dash['MT (HTA)']
    df_dash['Gaz_AO'], df_dash['Gaz_FSM'] = bp_ao.values, bp_fsm.values
    df_dash['BP (Total)'] = df_dash['Gaz_AO'] + df_dash['Gaz_FSM']
    df_dash['MP'] = mp_v.values
    df_dash['Total Gaz'] = df_dash['BP (Total)'] + df_dash['MP']
    df_dash.loc['TOTAL DD'] = df_dash.sum()

    cols_ca = ['Elec_AO', 'Elec_FSM', 'BT (Total)', 'MT (HTA)', 'Total Électricité', 'Gaz_AO', 'Gaz_FSM', 'BP (Total)', 'MP', 'Total Gaz']
    df_ca = pd.DataFrame(0.0, index=range(1, 13), columns=cols_ca)
    ttc_tabs = {cat: pd.DataFrame(0.0, index=range(1, 13), columns=['Ems Cycl', 'Ems HC', 'Annul', 'TOTAL']) for cat in ['HTA', 'MP', 'BT', 'BP']}
    
    configs_v2 = {
        'HTA': {'pos': (4, 6), 'cols': ['énergie active', 'Bonif', 'PMD', 'PMA', 'Redevances']},
        'MP':  {'pos': (5, 7), 'cols': ['énergie active', 'DMD', 'Redevances']},
        'BT':  {'pos': None,   'cols': ['Consom', 'Redev', 'Prest']},
        'BP':  {'pos': None,   'cols': ['Consom', 'Redev', 'Prest']}
    }

    files_v2 = glob.glob("*.xlsx") + glob.glob("data/*.xlsx")
    for file in files_v2:
        fname = os.path.basename(file).upper()
        cat = next((k for k in configs_v2 if k in fname), None)
        if not cat or 'RESULTAT' in fname: continue
        try:
            df = pd.read_excel(file, skiprows=5)
            df.columns = [str(c).strip() for c in df.columns]
            col_nat = [c for c in df.columns if 'NATURE' in c.upper()][0]
            col_ttc = [c for c in df.columns if 'TTC' in c.upper() or 'TOTAL' in c.upper()][-1]
            df[col_ttc] = pd.to_numeric(df[col_ttc], errors='coerce').fillna(0)
            if cat in ['BT', 'BP']:
                col_date = [c for c in df.columns if 'DATE' in c.upper()][0]
                df['Mois'] = pd.to_datetime(df[col_date]).dt.month
            else:
                col_fact = [c for c in df.columns if 'FACTURE' in c.upper()][0]
                s, e = configs_v2[cat]['pos']
                df['Mois'] = df[col_fact].astype(str).str[s:e].str.extract(r'(\d+)').astype(float).fillna(0).astype(int)
            
            target_cols = [c for c in df.columns if any(x.lower() in c.lower() for x in configs_v2[cat]['cols'])]
            df['Sum_Raw'] = df[target_cols].sum(axis=1)
            df['Net_Amount'] = df.apply(lambda x: -x['Sum_Raw'] if 'ANNUL' in str(x[col_nat]).upper() else x['Sum_Raw'], axis=1)

            if cat == 'BT':
                df_ca['Elec_AO'] += df[df.astype(str).apply(lambda x: x.str.contains('AO', case=False)).any(axis=1)].groupby('Mois')['Net_Amount'].sum()
                df_ca['Elec_FSM'] += df[df.astype(str).apply(lambda x: x.str.contains('FSM', case=False)).any(axis=1)].groupby('Mois')['Net_Amount'].sum()
            elif cat == 'BP':
                df_ca['Gaz_AO'] += df[df.astype(str).apply(lambda x: x.str.contains('AO', case=False)).any(axis=1)].groupby('Mois')['Net_Amount'].sum()
                df_ca['Gaz_FSM'] += df[df.astype(str).apply(lambda x: x.str.contains('FSM', case=False)).any(axis=1)].groupby('Mois')['Net_Amount'].sum()
            elif cat == 'HTA': df_ca['MT (HTA)'] += df.groupby('Mois')['Net_Amount'].sum()
            elif cat == 'MP': df_ca['MP'] += df.groupby('Mois')['Net_Amount'].sum()

            for m in range(1, 13):
                m_data = df[df['Mois'] == m]
                if m_data.empty: continue
                ttc_tabs[cat].at[m, 'Annul'] += m_data[m_data[col_nat].astype(str).str.contains('ANNUL', case=False, na=False)][col_ttc].sum()
                ttc_tabs[cat].at[m, 'Ems HC'] += m_data[m_data[col_nat].astype(str).str.contains('HC', case=False, na=False)][col_ttc].sum()
                ttc_tabs[cat].at[m, 'Ems Cycl'] += m_data[~m_data[col_nat].astype(str).str.contains('ANNUL|HC', case=False, na=False)][col_ttc].sum()
        except: continue

    df_ca['BT (Total)'] = df_ca['Elec_AO'] + df_ca['Elec_FSM']
    df_ca['Total Électricité'] = df_ca['BT (Total)'] + df_ca['MT (HTA)']
    df_ca['BP (Total)'] = df_ca['Gaz_AO'] + df_ca['Gaz_FSM']
    df_ca['Total Gaz'] = df_ca['BP (Total)'] + df_ca['MP']
    df_ca.index = months_fr
    df_ca.loc['TOTAL DD'] = df_ca.sum()
    df_ca_mil = df_ca / 1_000_000
    for cat in ttc_tabs:
        ttc_tabs[cat]['TOTAL'] = ttc_tabs[cat]['Ems Cycl'] + ttc_tabs[cat]['Ems HC'] - ttc_tabs[cat]['Annul']
        ttc_tabs[cat].index = months_fr
        ttc_tabs[cat].loc['TOTAL DD'] = ttc_tabs[cat].sum()

    output_file = "Rapport_Final_TB.xlsx"
    with pd.ExcelWriter(output_file, engine='openpyxl') as writer:
        
        # --- دالة التنسيق المحصور مع العنوان الموسط ---
        def apply_final_styling(ws, s_row, e_row, s_col, e_col, title_text):
            # دمج الخلايا للعنوان فوق الجدول مباشرة
            ws.merge_cells(start_row=s_row-1, start_column=s_col, end_row=s_row-1, end_column=e_col)
            title_cell = ws.cell(s_row-1, s_col, title_text)
            title_cell.font = Font(size=14, bold=True, color="2F5597")
            title_cell.alignment = Alignment(horizontal="center", vertical="center")
            
            for r in range(s_row, e_row + 1):
                for c in range(s_col, e_col + 1):
                    cell = ws.cell(r, c)
                    cell.alignment = Alignment(horizontal="center", vertical="center")
                    cell.border = std_border
                    if r == s_row: cell.fill, cell.font = header_fill, white_bold
                    if "TOTAL" in str(ws.cell(r, s_col).value).upper(): cell.fill, cell.font = total_fill, Font(bold=True)

        # --- ورقة 1: Dashboard ---
        df_dash.to_excel(writer, sheet_name='Dashboard', startrow=3)
        (df_dash/1000000).to_excel(writer, sheet_name='Dashboard', startrow=27)
        ws_dash = writer.sheets['Dashboard']
        ws_dash.sheet_view.showGridLines = False
        apply_final_styling(ws_dash, 4, 17, 1, 11, "RÉCAPITULATIF DES VENTES (UNITÉ: KWH)")
        apply_final_styling(ws_dash, 28, 41, 1, 11, "RÉCAPITULATIF DES VENTES (UNITÉ: GWH)")
        for r in range(4, 42): 
            for c in range(2, 12): ws_dash.cell(r, c).number_format = fmt_accounting

        # --- ورقة 2: Détail par Agence (كما هي) ---
        ws_det = writer.book.create_sheet("Détail par Agence")
        ws_det.sheet_view.showGridLines = False
        curr_row = 1
        for energy in energies:
            ws_det.merge_cells(start_row=curr_row, start_column=1, end_row=curr_row, end_column=53)
            ws_det.cell(curr_row, 1, f"DÉTAIL PAR NATURE : {energy}").font = Font(size=14, bold=True); ws_det.cell(curr_row, 1).alignment = Alignment(horizontal="center")
            curr_row += 2
            ws_det.cell(curr_row, 1, "AGENCES").fill = header_fill; ws_det.cell(curr_row, 1).font = white_bold
            for m_idx, col_name in enumerate(months_fr + ['TOTAL ANNUEL']):
                sc = 2 + (m_idx * 4)
                ws_det.merge_cells(start_row=curr_row, start_column=sc, end_row=curr_row, end_column=sc+3)
                c = ws_det.cell(curr_row, sc, col_name); c.fill = mid_fill; c.font = white_bold; c.alignment = Alignment(horizontal="center")
                for i, sh in enumerate(['CYCL', 'HC', 'ANNUL', 'TOTAL']):
                    sub = ws_det.cell(curr_row+1, sc+i, sh); sub.fill = header_fill; sub.font = white_bold; sub.alignment = Alignment(horizontal="center")
            df_e = data_v1[energy]; curr_row += 2
            ag_totals = {code: {'CYCL':0, 'HC':0, 'ANNUL':0} for code in all_codes}
            for code in all_codes:
                ws_det.cell(curr_row, 1, code_to_name.get(code, code)).border = std_border
                for m in range(1, 13):
                    sc = 2 + (m-1)*4
                    d = df_e[(df_e['Ag_Code']==code) & (df_e['Mois']==m)]
                    v_c, v_h, v_a = d[d['Nat']=='CYCL']['Val'].sum(), d[d['Nat']=='HC']['Val'].sum(), d[d['Nat']=='ANNUL']['Val'].sum()
                    vals = [v_c, v_h, v_a, v_c + v_h - v_a]
                    for i, v in enumerate(vals):
                        cell = ws_det.cell(curr_row, sc+i, v); cell.border = std_border; cell.number_format = fmt_accounting
                        if i < 3: ag_totals[code][['CYCL','HC','ANNUL'][i]] += v
                sc_ann = 2 + 12*4
                ann_vals = [ag_totals[code]['CYCL'], ag_totals[code]['HC'], ag_totals[code]['ANNUL'], ag_totals[code]['CYCL']+ag_totals[code]['HC']-ag_totals[code]['ANNUL']]
                for i, v in enumerate(ann_vals):
                    cell = ws_det.cell(curr_row, sc_ann+i, v); cell.border = std_border; cell.font = Font(bold=True); cell.fill = total_fill; cell.number_format = fmt_accounting
                curr_row += 1
            ws_det.cell(curr_row, 1, "TOTAL DD").font = Font(bold=True); ws_det.cell(curr_row, 1).fill = total_fill; ws_det.cell(curr_row, 1).border = std_border
            for m in range(1, 14):
                sc = 2 + (m-1)*4
                for i, nat in enumerate(['CYCL', 'HC', 'ANNUL', 'TOTAL']):
                    if m <= 12:
                        d_m = df_e[df_e['Mois']==m]
                        tv = d_m[d_m['Nat']=='CYCL']['Val'].sum() + d_m[d_m['Nat']=='HC']['Val'].sum() - d_m[d_m['Nat']=='ANNUL']['Val'].sum() if nat == 'TOTAL' else d_m[d_m['Nat']==nat]['Val'].sum()
                    else:
                        tv = sum(ag_totals[c]['CYCL'] + ag_totals[c]['HC'] - ag_totals[c]['ANNUL'] for c in all_codes) if nat == 'TOTAL' else sum(ag_totals[c][nat] for c in all_codes)
                    cell = ws_det.cell(curr_row, sc+i, tv); cell.border = std_border; cell.font = Font(bold=True); cell.fill = total_fill; cell.number_format = fmt_accounting
            curr_row += 6

        # --- ورقة 3: CA_Dashboard ---
        df_ca.to_excel(writer, sheet_name='CA_Dashboard', startrow=2)
        df_ca_mil.to_excel(writer, sheet_name='CA_Dashboard', startrow=19)
        ws_ca = writer.sheets['CA_Dashboard']
        ws_ca.sheet_view.showGridLines = False
        apply_final_styling(ws_ca, 3, 16, 1, 11, "TABLEAU CHIFFRE D'AFFAIRES - RÉEL (DA)")
        apply_final_styling(ws_ca, 20, 33, 1, 11, "TABLEAU CHIFFRE D'AFFAIRES - EN MILLIONS (MDA)")
        for r in range(3, 34):
            for c in range(2, 12): ws_ca.cell(r, c).number_format = fmt_accounting

        # --- ورقة 4: Details TTC (كما هي) ---
        ws_ttc = writer.book.create_sheet('Details TTC')
        ws_ttc.sheet_view.showGridLines = False
        start_r = 0
        for cat in ['HTA', 'MP', 'BT', 'BP']:
            ws_ttc.cell(row=start_r+1, column=1, value=f"TABLEAU TTC - {cat} (CYCL + HC - ANNUL)").font = Font(bold=True, color="2F5597")
            ttc_tabs[cat].to_excel(writer, sheet_name='Details TTC', startrow=start_r+1)
            for r in range(start_r+2, start_r+16):
                for c in range(1, 6):
                    cell = ws_ttc.cell(r, c); cell.border = std_border; cell.alignment = Alignment(horizontal="center")
                    if r == start_r+2 and c > 1: cell.fill, cell.font = header_fill, white_bold
                    if "TOTAL DD" in str(ws_ttc.cell(r, 1).value): cell.fill, cell.font = total_fill, Font(bold=True)
                    if isinstance(cell.value, (int, float)): cell.number_format = fmt_accounting
            start_r += 17

        # --- أوراق Trimestres (كما هي) ---
        df_bt_v3, df_bp_v3 = get_stats_data_v3('BT'), get_stats_data_v3('BP')
        for t_num, m_list in [(1,[1,2,3]), (2,[4,5,6]), (3,[7,8,9]), (4,[10,11,12])]:
            sheet_name = f"Trimestre {t_num}"
            ws_st = writer.book.create_sheet(sheet_name); ws_st.sheet_view.showGridLines = False
            s_row = 2
            for lab, df_e in [("ÉLECTRICITÉ (BT)", df_bt_v3), ("GAZ (BP)", df_bp_v3)]:
                ws_st.merge_cells(start_row=s_row, start_column=1, end_row=s_row, end_column=38)
                ws_st.cell(s_row, 1, lab).font = Font(size=14, bold=True); ws_st.cell(s_row, 1).alignment = Alignment(horizontal="center")
                r1, r2, r3 = s_row+1, s_row+2, s_row+3
                ws_st.cell(r1, 1, "Année").fill = mid_fill; ws_st.cell(r1, 1).font = white_bold
                for idx, m_idx in enumerate(m_list + ["TOTAL TRIMESTRE"]):
                    c_s = 2 + (idx * 9)
                    ws_st.merge_cells(start_row=r1, start_column=c_s, end_row=r1, end_column=c_s+8)
                    ws_st.cell(r1, c_s, months_fr[m_idx-1] if isinstance(m_idx, int) else m_idx).fill = mid_fill; ws_st.cell(r1, c_s).font = white_bold; ws_st.cell(r1, c_s).alignment = Alignment(horizontal="center")
                    for j, cat_l in enumerate(["AO", "FSM", "BT/BP TOTAL"]):
                        c_sub = c_s + (j*3)
                        ws_st.merge_cells(start_row=r2, start_column=c_sub, end_row=r2, end_column=c_sub+2)
                        ws_st.cell(r2, c_sub, cat_l).fill = header_fill; ws_st.cell(r2, c_sub).font = white_bold; ws_st.cell(r2, c_sub).alignment = Alignment(horizontal="center")
                        for k, sub_l in enumerate(["Nulles", "Total", "Taux (%)"]):
                            cell = ws_st.cell(r3, c_sub+k, sub_l); cell.fill = PatternFill(start_color="8EA9DB", fill_type="solid"); cell.font = Font(bold=True); cell.alignment = Alignment(horizontal="center"); cell.border = std_border
                r4 = r3 + 1
                ws_st.cell(r4, 1, str(df_e['Annee'].iloc[0]) if not df_e.empty else "-").border = std_border; ws_st.cell(r4, 1).alignment = Alignment(horizontal="center")
                for idx, m_idx in enumerate(m_list + ["TOTAL"]):
                    m_data = df_e[df_e['Mois'].isin(m_list)] if m_idx == "TOTAL" else df_e[df_e['Mois'] == m_idx]
                    off = 2 + (idx * 9)
                    for j, cat in enumerate(["AO", "FSM", "Total"]):
                        c_df = m_data[m_data['Cat'] == cat] if cat != "Total" else m_data[m_data['Cat'].isin(['AO', 'FSM'])]
                        n, t = c_df['Is_Nulle'].sum(), len(c_df)
                        c_n, c_t, c_x = ws_st.cell(r4, off+(j*3), n), ws_st.cell(r4, off+(j*3)+1, t), ws_st.cell(r4, off+(j*3)+2, (n/t if t>0 else 0))
                        for c in [c_n, c_t, c_x]: 
                            c.border = std_border; c.alignment = Alignment(horizontal="center")
                            if m_idx == "TOTAL": c.fill = total_fill; c.font = Font(bold=True)
                        c_x.number_format = '0.00%'
                s_row = r4 + 4
            for i in range(1, 40): ws_st.column_dimensions[get_column_letter(i)].width = 13

        for sn in writer.sheets:
            ws = writer.sheets[sn]
            for i in range(1, 55): ws.column_dimensions[get_column_letter(i)].width = 22

    print(f"✅ مريقل يا يوسف! تم تجميد التنسيق النهائي: {output_file}")

if __name__ == "__main__":
    run_mega_process()
